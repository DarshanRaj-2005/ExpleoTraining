import openpyxl

def get_data(filename, sheet_name):
    data = []
    workbook = openpyxl.load_workbook("./data_files/"+filename)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_cols+1):
            value = sheet.cell(r, c).value
            if value == None:
                row_list.append("")
            else:
                row_list.append(value)
        data.append(row_list)

    return data